import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADERS = ["SKU", "Артикул WB (nm_id)", "Баркод"]

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")


def load_sku_table(conn: sqlite3.Connection, run_date: str | None = None) -> list[dict]:
    """Return [{sku, nm_id, barcode}, ...] from fbo_stocks for the given run_date.

    Uses nm_id → canonical SKU mapping from fbo_sales (same priority as load_plan_inputs).
    One row per unique (sku, nm_id, barcode) triple, sorted by sku then barcode.
    """
    if run_date is None:
        row = conn.execute("SELECT MAX(run_date) FROM fbo_stocks").fetchone()
        run_date = row[0] if row and row[0] else None
    if not run_date:
        return []

    nm_to_sku: dict[int, str] = {}
    for row in conn.execute(
        "SELECT DISTINCT nm_id, supplier_article FROM fbo_sales "
        "WHERE nm_id IS NOT NULL AND supplier_article IS NOT NULL"
    ):
        nm_id = int(row[0])
        if nm_id not in nm_to_sku:
            nm_to_sku[nm_id] = row[1]

    seen: set[tuple] = set()
    rows: list[dict] = []
    for row in conn.execute(
        """SELECT nm_id, vendor_code,
                  json_extract(raw_payload, '$.barcode') AS barcode
           FROM fbo_stocks
           WHERE run_date = ?
           GROUP BY nm_id, barcode""",
        (run_date,),
    ):
        nm_id_val = int(row[0] or 0)
        vc, bc = row[1], row[2]
        sku = nm_to_sku.get(nm_id_val) or vc
        if not sku:
            continue
        bc_str = str(bc) if bc else ""
        key = (sku, nm_id_val, bc_str)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"sku": sku, "nm_id": nm_id_val or "", "barcode": bc_str})

    return sorted(rows, key=lambda r: (r["sku"], r["barcode"]))


def write_sku_db(conn: sqlite3.Connection, run_date: str | None, output_path: Path) -> Path:
    """Generate SKU database Excel: SKU | nm_id | Barcode.

    Returns path to the created file.
    """
    rows = load_sku_table(conn, run_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "WB SKU Database"

    bold = Font(bold=True)

    header_row = ws.append(HEADERS) or ws.max_row
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = bold
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r["sku"], r["nm_id"] or "", r["barcode"]])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
