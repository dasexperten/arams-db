from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .calc import ZONE_DEFICIT, ZONE_NORMAL, ZONE_OVERSTOCK

# Кластер | Склад | SKU | Остаток | Продажи | К | Зона | Поставка | Примечание
HEADERS = ["Кластер", "Склад", "SKU", "Остаток", "Продажи", "К", "Зона", "Поставка", "Примечание"]

_ZONE_FILL = {
    ZONE_DEFICIT: PatternFill("solid", fgColor="FCEBEB"),
    ZONE_NORMAL: PatternFill("solid", fgColor="EAF3DE"),
    ZONE_OVERSTOCK: PatternFill("solid", fgColor="F1EFE8"),
}

_COL_H = 8   # Поставка (1-indexed)
_COL_G = 7   # Зона


def write_excel(plans: list[dict], run_date: str, output_dir: Path) -> Path:
    """Generate Excel supply plan. Returns path to created file.

    Columns: A=Кластер B=Склад C=SKU D=Остаток E=Продажи F=К G=Зона H=Поставка I=Примечание
    Sorted: clusters by descending total to_ship, within cluster warehouses alpha,
    within warehouse to_ship DESC then SKU alpha.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"wb_fbo_{run_date}.xlsx"

    oos_count = sum(1 for p in plans if "🔴 Товар вышел" in (p.get("flag") or ""))

    # Compute cluster totals for sorting
    cluster_totals: dict[str, int] = {}
    for p in plans:
        cl = p.get("cluster") or ""
        cluster_totals[cl] = cluster_totals.get(cl, 0) + (p.get("to_ship") or 0)

    def sort_key(p):
        cl = p.get("cluster") or ""
        wh = p.get("warehouse") or ""
        is_oos = 1 if "🔴 Товар вышел" in (p.get("flag") or "") else 0
        return (-cluster_totals.get(cl, 0), cl, wh, -is_oos, -(p.get("to_ship") or 0), p.get("sku") or "")

    sorted_plans = sorted(plans, key=sort_key)

    # Build nested structure: cluster → warehouse → [rows]
    clusters: list[str] = []
    cluster_warehouses: dict[str, list[str]] = {}
    cluster_wh_rows: dict[tuple, list[dict]] = {}
    for p in sorted_plans:
        cl = p.get("cluster") or ""
        wh = p.get("warehouse") or ""
        if cl not in cluster_warehouses:
            clusters.append(cl)
            cluster_warehouses[cl] = []
        if wh not in cluster_warehouses[cl]:
            cluster_warehouses[cl].append(wh)
        key = (cl, wh)
        if key not in cluster_wh_rows:
            cluster_wh_rows[key] = []
        cluster_wh_rows[key].append(p)

    wb = Workbook()
    ws = wb.active
    ws.title = f"WB-FBO {run_date}"

    bold = Font(bold=True)
    cluster_fill = PatternFill("solid", fgColor="D9D9D9")

    if oos_count:
        ws.append([f"⚠️ {oos_count} позиций в out-of-stock — приоритет отгрузки"] + [""] * 8)
        ws.merge_cells("A1:I1")
        ws["A1"].font = bold
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    ws.append(HEADERS)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = bold

    for cl in clusters:
        cl_total_sales = 0
        cl_total_ship = 0
        cl_k_num = 0.0
        cl_k_den = 0

        for wh in cluster_warehouses[cl]:
            items = cluster_wh_rows[(cl, wh)]
            wh_total_sales = 0
            wh_total_ship = 0
            wh_k_num = 0.0
            wh_k_den = 0

            for item in items:
                sku = item.get("sku") or ""
                stock = item.get("stock") or 0
                sales = item.get("sales_30d") or 0
                k = item.get("k")
                zone = item.get("zone") or ZONE_NORMAL
                to_ship = item.get("to_ship") or 0
                flag = item.get("flag") or ""

                k_display = f"{k:.2f}" if k is not None else "—"
                ws.append([cl, wh, sku, stock, sales, k_display, zone, to_ship, flag])
                cur = ws.max_row

                ws.cell(row=cur, column=_COL_H).font = bold

                fill = _ZONE_FILL.get(zone)
                if fill:
                    ws.cell(row=cur, column=_COL_G).fill = fill

                wh_total_sales += sales
                wh_total_ship += to_ship
                cl_total_sales += sales
                cl_total_ship += to_ship
                if k is not None and sales > 0:
                    wh_k_num += k * sales
                    wh_k_den += sales
                    cl_k_num += k * sales
                    cl_k_den += sales

            wh_avg_k = (wh_k_num / wh_k_den) if wh_k_den > 0 else None
            wh_avg_k_display = f"{wh_avg_k:.2f}" if wh_avg_k is not None else "—"
            ws.append([cl, f"ИТОГО {wh.upper()}", "", "", wh_total_sales,
                       wh_avg_k_display, "", wh_total_ship, ""])
            summary_row = ws.max_row
            for cell in ws[summary_row]:
                cell.font = bold

        cl_avg_k = (cl_k_num / cl_k_den) if cl_k_den > 0 else None
        cl_avg_k_display = f"{cl_avg_k:.2f}" if cl_avg_k is not None else "—"
        ws.append([f"ИТОГО КЛАСТЕР {cl.upper()}", "", "", "", cl_total_sales,
                   cl_avg_k_display, "", cl_total_ship, ""])
        cl_row = ws.max_row
        for cell in ws[cl_row]:
            cell.font = bold
            cell.fill = cluster_fill

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 46

    wb.save(str(out_path))
    return out_path


def build_summary(plans: list[dict]) -> dict:
    """Build stats dict for Telegram caption."""
    total = len(plans)
    to_ship_count = sum(1 for p in plans if (p.get("to_ship") or 0) > 0)
    to_ship_units = sum(p.get("to_ship") or 0 for p in plans)
    normal = sum(1 for p in plans if p.get("zone") == ZONE_NORMAL and not (p.get("to_ship") or 0) > 0)
    overstock = sum(1 for p in plans if p.get("zone") == ZONE_OVERSTOCK)
    oos = sum(1 for p in plans if "🔴 Товар вышел" in (p.get("flag") or ""))
    unknown_pack = sum(1 for p in plans if "Unknown pack" in (p.get("flag") or ""))
    unknown_wh = sum(1 for p in plans if "Unknown warehouse" in (p.get("flag") or ""))

    cluster_ship: dict[str, int] = {}
    for p in plans:
        cl = p.get("cluster") or "?"
        cluster_ship[cl] = cluster_ship.get(cl, 0) + (p.get("to_ship") or 0)

    return {
        "total": total,
        "to_ship_count": to_ship_count,
        "to_ship_units": to_ship_units,
        "normal": normal,
        "overstock": overstock,
        "oos": oos,
        "unknown_pack": unknown_pack,
        "unknown_wh": unknown_wh,
        "cluster_ship": cluster_ship,
    }
