from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .calc import ZONE_DEFICIT, ZONE_NORMAL, ZONE_OVERSTOCK
from .clusters import CLUSTER_ORDER, CLUSTER_CODES

HEADERS = ["SKU", "Остаток", "Продажи", "К", "Зона", "Поставка", "Примечание"]

_ZONE_FILL = {
    ZONE_DEFICIT:   PatternFill("solid", fgColor="FCEBEB"),
    ZONE_NORMAL:    PatternFill("solid", fgColor="EAF3DE"),
    ZONE_OVERSTOCK: PatternFill("solid", fgColor="F1EFE8"),
}

_COL_SHIP = 6
_COL_ZONE = 5


def _date_str(run_date: str) -> str:
    """Convert YYYY-MM-DD to ddmmyy."""
    return datetime.strptime(run_date, "%Y-%m-%d").strftime("%d%m%y")


def write_excel(plans: list[dict], run_date: str, output_dir: Path) -> list[Path]:
    """Generate one Excel file per cluster. Returns list of created paths.

    Filename format: ozon_{CODE}_{ddmmyy}.xlsx  (e.g. ozon_MSC_240426.xlsx)
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_sfx = _date_str(run_date)

    by_cluster: dict[str, list[dict]] = {}
    for p in plans:
        cl = p.get("cluster") or "UNKNOWN"
        by_cluster.setdefault(cl, []).append(p)

    def cl_order(cl):
        return CLUSTER_ORDER.index(cl) if cl in CLUSTER_ORDER else len(CLUSTER_ORDER)

    out_paths: list[Path] = []
    for cl in sorted(by_cluster, key=cl_order):
        code = CLUSTER_CODES.get(cl, cl[:3].upper())
        out_path = output_dir / f"ozon_{code}_{date_sfx}.xlsx"
        _write_cluster_file(by_cluster[cl], cl, run_date, out_path)
        out_paths.append(out_path)

    return out_paths


def _write_cluster_file(plans: list[dict], cluster: str, run_date: str, out_path: Path) -> None:
    oos_count = sum(1 for p in plans if "🔴 Товар вышел" in (p.get("flag") or ""))

    def sort_key(p):
        sales = p.get("sales_30d") or 0
        return (1 if sales == 0 else 0, -sales, p.get("sku") or "")

    sorted_plans = sorted(plans, key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = cluster

    bold = Font(bold=True)

    if oos_count:
        ws.append([f"⚠️ {oos_count} позиций в out-of-stock — приоритет отгрузки"] + [""] * 6)
        ws.merge_cells("A1:G1")
        ws["A1"].font = bold
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    ws.append(HEADERS)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = bold

    total_sales = 0
    total_ship = 0
    k_num = 0.0
    k_den = 0

    for item in sorted_plans:
        sku = item.get("sku") or ""
        stock = item.get("stock") or 0
        sales = item.get("sales_30d") or 0
        k = item.get("k")
        zone = item.get("zone") or ZONE_NORMAL
        to_ship = item.get("to_ship") or 0
        flag = item.get("flag") or ""
        global_oos = item.get("global_oos", False)

        k_display = f"{k:.2f}" if k is not None else "—"
        to_ship_display = f"{to_ship} ⚠️" if global_oos and to_ship > 0 else to_ship
        ws.append([sku, stock, sales, k_display, zone, to_ship_display, flag])
        cur = ws.max_row
        ws.cell(row=cur, column=_COL_SHIP).font = bold
        fill = _ZONE_FILL.get(zone)
        if fill:
            ws.cell(row=cur, column=_COL_ZONE).fill = fill

        total_sales += sales
        total_ship += to_ship
        if k is not None and sales > 0:
            k_num += k * sales
            k_den += sales

    avg_k = (k_num / k_den) if k_den > 0 else None
    avg_k_display = f"{avg_k:.2f}" if avg_k is not None else "—"
    cluster_fill = PatternFill("solid", fgColor="D9D9D9")
    ws.append(["ИТОГО", "", total_sales, avg_k_display, "", total_ship, ""])
    tot_row = ws.max_row
    for cell in ws[tot_row]:
        cell.font = bold
        cell.fill = cluster_fill

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 46

    # Sheet 2: supply order — артикул | имя | количество
    ship_totals: dict[str, dict] = {}
    for p in plans:
        qty = p.get("to_ship") or 0
        if qty <= 0:
            continue
        sku = p.get("sku") or ""
        if sku not in ship_totals:
            ship_totals[sku] = {"item_name": p.get("item_name") or "", "qty": 0}
        ship_totals[sku]["qty"] += qty

    if ship_totals:
        ws2 = wb.create_sheet("К отгрузке")
        ws2.append(["Артикул", "Название", "Количество"])
        hdr2 = ws2.max_row
        for cell in ws2[hdr2]:
            cell.font = bold
        for sku in sorted(ship_totals):
            ws2.append([sku, ship_totals[sku]["item_name"], ship_totals[sku]["qty"]])
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 14

    wb.save(str(out_path))


def build_summary(plans: list[dict]) -> dict:
    to_ship_count = sum(1 for p in plans if (p.get("to_ship") or 0) > 0)
    to_ship_units = sum(p.get("to_ship") or 0 for p in plans)
    normal = sum(1 for p in plans if p.get("zone") == ZONE_NORMAL and not (p.get("to_ship") or 0) > 0)
    overstock = sum(1 for p in plans if p.get("zone") == ZONE_OVERSTOCK)
    oos = sum(1 for p in plans if p.get("global_oos"))
    unknown_pack = sum(1 for p in plans if "Unknown pack" in (p.get("flag") or ""))

    cluster_ship: dict[str, int] = {}
    for p in plans:
        cl = p.get("cluster") or "?"
        cluster_ship[cl] = cluster_ship.get(cl, 0) + (p.get("to_ship") or 0)

    return {
        "total": len(plans),
        "to_ship_count": to_ship_count,
        "to_ship_units": to_ship_units,
        "normal": normal,
        "overstock": overstock,
        "oos": oos,
        "unknown_pack": unknown_pack,
        "cluster_ship": cluster_ship,
    }
