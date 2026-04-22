import argparse
import csv
import json
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv

from ozon_perf import OzonPerformanceClient
from ozon_perf import analyze, dashboard, db, etl
from ozon_seller import OzonSellerClient, SellerAPI
from ozon_seller import db as seller_db
from ozon_seller import etl as seller_etl
from wb_seller import WBSellerClient, WBFeedbacksAPI
from wb_seller import db as wb_db
from wb_seller import etl as wb_etl


def _load_catalog() -> dict[str, str]:
    """Build {ozon_sku: name} + {offer_id: name} from products.csv + Excel + DB cache."""
    import re

    offer_names: dict[str, str] = {}
    csv_path = Path(__file__).parent / "data" / "products.csv"
    if csv_path.exists():
        with csv_path.open(encoding="utf-8") as f:
            offer_names = {row["offer_id"]: row["name"]
                          for row in csv.DictReader(f) if row.get("offer_id")}

    catalog: dict[str, str] = dict(offer_names)

    # Read Ozon SKU → offer_id mapping from the Excel flat file.
    xlsx_path = Path(__file__).parent / "ozon_seller" / "Ozon Products.xlsx"
    if xlsx_path.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    if row and str(row[0] or "").strip() == "Артикул":
                        header_found = True
                    continue
                raw_offer = str(row[0] or "").lstrip("'").strip()
                sku_val = row[2]
                if not raw_offer or not sku_val:
                    continue
                base_offer = re.sub(r"A+$", "", raw_offer)
                ozon_sku = str(int(sku_val))
                name = offer_names.get(base_offer)
                if name:
                    catalog[ozon_sku] = name
            wb.close()
        except Exception:
            pass

    # Also merge any additional entries from the DB cache.
    try:
        with seller_db.connect() as conn:
            for ozon_sku, offer_id in seller_db.sku_to_offer_id(conn).items():
                if ozon_sku not in catalog:
                    name = offer_names.get(offer_id)
                    if name:
                        catalog[ozon_sku] = name
    except Exception:
        pass

    return catalog


def _sku_label(catalog: dict[str, str], sku: str) -> str:
    return catalog.get(str(sku), str(sku)) if sku else ""


def _extract_author(obj: dict) -> str:
    """Ozon returns author as a nested dict {first_name, last_name} or as a plain string."""
    raw = (
        obj.get("author_name")
        or obj.get("author")
        or obj.get("name")
        or ""
    )
    if isinstance(raw, dict):
        return (raw.get("first_name") or "").strip()
    return str(raw).strip()


def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def cmd_init(_: argparse.Namespace) -> int:
    db.init_schema()
    seller_db.init_schema()
    wb_db.init_schema()
    print("perf schema ready:  ", db._db_path())
    print("seller schema ready:", seller_db._db_path())
    print("wb schema ready:    ", wb_db._db_path())
    return 0


def cmd_ping_seller(_: argparse.Namespace) -> int:
    with OzonSellerClient() as c:
        data = SellerAPI(c).reviews_count()
    print("seller auth ok, review counts:")
    print(json.dumps(data, indent=2, ensure_ascii=False))
    return 0


def cmd_debug_review(_: argparse.Namespace) -> int:
    """Dump raw JSON of first review to stdout — used to inspect Ozon field names."""
    with OzonSellerClient() as c:
        page = SellerAPI(c).reviews_list(status="ALL", limit=1)
    reviews = page.get("reviews") or []
    if not reviews:
        print("No reviews returned")
        return 0
    print(json.dumps(reviews[0], indent=2, ensure_ascii=False))
    return 0


def cmd_ping_questions(_: argparse.Namespace) -> int:
    with OzonSellerClient() as c:
        api = SellerAPI(c)
        counts = api.questions_count()
        print("questions API ok, counts:")
        print(json.dumps(counts, indent=2, ensure_ascii=False))
        for status in ("UNPROCESSED", "PROCESSED", "ALL"):
            page = api.questions_list(status=status, limit=1)
            questions = page.get("questions") or []
            if questions:
                q = questions[0]
                print(f"\nfirst question (status={status}) fields: {list(q.keys())}")
                print(json.dumps(q, indent=2, ensure_ascii=False, default=str))
                break
        else:
            print("\n(no questions found in any status)")
    return 0


def cmd_sync_product_skus(_: argparse.Namespace) -> int:
    """Parse Ozon Products.xlsx and save ozon_sku → offer_id mapping to DB."""
    import re
    import openpyxl

    xlsx_path = Path(__file__).parent / "ozon_seller" / "Ozon Products.xlsx"
    if not xlsx_path.exists():
        print(f"error: {xlsx_path} not found", file=sys.stderr)
        return 1

    seller_db.init_schema()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0] or "").strip() == "Артикул":
                header_found = True
            continue
        raw_offer = str(row[0] or "").lstrip("'").strip()
        sku_val = row[2]
        if not raw_offer or not sku_val:
            continue
        base_offer = re.sub(r"A+$", "", raw_offer)
        ozon_sku = str(int(sku_val))
        rows.append({"ozon_sku": ozon_sku, "offer_id": base_offer})
    wb.close()

    print(f"parsed {len(rows)} rows from {xlsx_path.name}")

    if rows:
        with seller_db.connect() as conn:
            n = seller_db.upsert_product_skus(conn, rows)
        print(f"saved {n} ozon_sku → offer_id mappings to DB")

    print("\nFull mapping (ozon_sku → offer_id):")
    with seller_db.connect() as conn:
        db_rows = conn.execute(
            "SELECT ozon_sku, offer_id, synced_at FROM product_skus ORDER BY offer_id"
        ).fetchall()
    for row in db_rows:
        print(f"  {row['ozon_sku']:>15}  →  {row['offer_id']}")

    return 0


def cmd_sync_reviews(args: argparse.Namespace) -> int:
    seller_db.init_schema()
    with OzonSellerClient() as c:
        result = seller_etl.sync_reviews(
            c,
            status=args.status,
            max_reviews=args.max,
            with_comments=args.with_comments,
        )
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


def cmd_list_recent(args: argparse.Namespace) -> int:
    with OzonSellerClient() as c:
        page = SellerAPI(c).reviews_list(
            status=args.status, limit=args.count, sort_dir="DESC",
        )
    reviews = page.get("reviews") or []
    if not reviews:
        print(f"(no reviews found with status={args.status})")
        return 0
    print(f"{len(reviews)} most recent reviews (status={args.status}):")
    print("=" * 90)
    for r in reviews:
        rid = r.get("id", "")
        rating = r.get("rating", "?")
        sku = r.get("sku", "")
        author = _extract_author(r) or "(anonymous)"
        published = (r.get("published_at") or "")[:19]
        text = (r.get("text") or "").strip().replace("\n", " ")
        preview = text[:120] + ("…" if len(text) > 120 else "")
        if not preview:
            preview = "(no text, rating-only review)"
        print(f"id: {rid}")
        print(f"    ⭐ {rating}/5   SKU: {sku}   by: {author}   on: {published}")
        print(f"    “{preview}”")
        print("-" * 90)
    print()
    print("Pick an id and run: python cli.py draft-reply <id>")
    return 0


def cmd_draft_reply(args: argparse.Namespace) -> int:
    from ozon_seller.replier import draft_reply
    with OzonSellerClient() as c:
        info = SellerAPI(c).review_info(args.review_id)
    review = info.get("result") if isinstance(info.get("result"), dict) else info
    if not review or "id" not in review:
        review = {**(review or {}), "id": args.review_id}
    print("=" * 70)
    print(f"Review {args.review_id}")
    print("=" * 70)
    print(f"Author: {review.get('author') or review.get('name') or '(anonymous)'}")
    print(f"Rating: {review.get('rating')}/5   SKU: {review.get('sku')}")
    print(f"Published: {review.get('published_at')}")
    print("-" * 70)
    print(review.get("text") or "(empty review text)")
    print("=" * 70)
    print()
    draft = draft_reply(review)
    print("DRAFT REPLY:")
    print("-" * 70)
    print(draft.text)
    print("-" * 70)
    print(
        f"[{draft.model}] in={draft.input_tokens} out={draft.output_tokens} "
        f"cache_read={draft.cache_read_input_tokens} cache_write={draft.cache_creation_input_tokens} "
        f"stop={draft.stop_reason} chars={len(draft.text)}"
    )
    print()
    print("To post this draft as-is, run:")
    print(f"  python cli.py post-reply {args.review_id} \"<paste approved text here>\"")
    return 0


def cmd_post_reply(args: argparse.Namespace) -> int:
    if not args.text.strip():
        print("error: empty reply text", file=sys.stderr)
        return 1
    if args.confirm != "YES":
        print("refusing to post without --confirm YES (safety guard)", file=sys.stderr)
        print("the reply is public and cannot be edited after sending.", file=sys.stderr)
        return 1
    with OzonSellerClient() as c:
        result = SellerAPI(c).comment_create(
            review_id=args.review_id,
            text=args.text,
            mark_review_as_processed=not args.keep_unprocessed,
        )
    print("posted:")
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


def cmd_auto_reply(args: argparse.Namespace) -> int:
    """Reply to UNPROCESSED reviews one at a time: stream through them,
    skip (PROCESSED-mark) reviews with no text, and for each review with
    text — draft via Claude and post the reply (Ozon marks PROCESSED
    atomically). Stops after --max-replies successfully posted replies.

    Default: 1 reply per run, so hourly cron = 1 answer per hour.
    To clear a backlog, dispatch manually with a larger max-replies.
    """
    from ozon_seller.replier import draft_reply

    max_replies = max(1, int(args.max_replies))
    seller_db.init_schema()
    catalog = _load_catalog()

    replied: list[dict] = []
    no_text_marked: list[str] = []
    errors: list[dict] = []

    with OzonSellerClient() as c:
        api = SellerAPI(c)

        for review in api.reviews_iter(status="UNPROCESSED"):
            if len(replied) >= max_replies:
                break

            review_id = str(review.get("id") or "").strip()
            if not review_id:
                errors.append({"stage": "validate", "error": "review missing id"})
                continue

            text = (review.get("text") or "").strip()
            if not text:
                try:
                    api.change_status([review_id], "PROCESSED")
                    no_text_marked.append(review_id)
                except Exception as e:
                    errors.append({"review_id": review_id, "stage": "mark_no_text",
                                   "error": str(e)})
                continue

            sku = str(review.get("sku") or "")
            if sku and sku not in catalog:
                try:
                    items = api.product_info_list([sku])
                    new_rows = []
                    for item in items:
                        ozon_sku = str(item.get("sku") or item.get("id") or "")
                        offer_id = str(item.get("offer_id") or "")
                        if ozon_sku and offer_id:
                            new_rows.append({"ozon_sku": ozon_sku, "offer_id": offer_id})
                            name = catalog.get(offer_id)
                            if name:
                                catalog[ozon_sku] = name
                    if new_rows:
                        with seller_db.connect() as _conn:
                            seller_db.upsert_product_skus(_conn, new_rows)
                except Exception:
                    pass
            review = {**review, "product_name": _sku_label(catalog, sku)}

            try:
                draft = draft_reply(review)
            except Exception as e:
                errors.append({"review_id": review_id, "stage": "draft",
                               "error": str(e)})
                continue

            draft_text = (draft.text or "").strip()
            if not draft_text:
                errors.append({"review_id": review_id, "stage": "draft",
                               "error": "empty draft text"})
                continue

            try:
                api.comment_create(
                    review_id=review_id,
                    text=draft_text,
                    mark_review_as_processed=True,
                )
                replied.append({
                    "review_id": review_id,
                    "chars": len(draft_text),
                    "question": text,
                    "answer": draft_text,
                    "rating": review.get("rating"),
                    "author": _extract_author(review),
                    "sku": review.get("sku"),
                    "product_name": review.get("product_name") or "",
                    "published_at": review.get("published_at") or "",
                })
            except Exception as e:
                errors.append({"review_id": review_id, "stage": "post",
                               "error": str(e)})

    summary = {
        "status": "ok" if not errors else "partial",
        "max_replies": max_replies,
        "replied": len(replied),
        "no_text_marked": len(no_text_marked),
        "errors": len(errors),
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    if errors:
        print("errors:")
        print(json.dumps(errors, indent=2, ensure_ascii=False))

    _telegram_autoreply(summary, errors=errors, replied=replied)

    return 0 if not errors else 1


def cmd_auto_answer_questions(args: argparse.Namespace) -> int:
    """Stream UNPROCESSED questions, draft answer via Claude, post to Ozon.

    Default: 1 answer per run, so every-30min cron = up to 48 per day.
    To clear a backlog, dispatch manually with a larger --max-answers.
    """
    from ozon_seller.question_answerer import draft_answer

    max_answers = max(1, int(args.max_answers))
    max_inspect = max_answers * 50  # hard stop: never loop through more than this
    catalog = _load_catalog()

    answered: list[dict] = []
    errors: list[dict] = []
    inspected = 0

    with OzonSellerClient() as c:
        api = SellerAPI(c)

        for question in api.questions_iter(status="UNPROCESSED"):
            if len(answered) >= max_answers:
                break
            if inspected >= max_inspect:
                print(f"(reached max_inspect={max_inspect}, stopping)", flush=True)
                break
            inspected += 1

            question_id = str(
                question.get("id") or question.get("question_id") or ""
            ).strip()
            if not question_id:
                errors.append({"stage": "validate", "error": "question missing id"})
                continue

            text = (
                question.get("question_text") or question.get("text") or ""
            ).strip()
            if not text:
                print(f"  skip {question_id}: no text", flush=True)
                continue

            sku = str(question.get("sku_id") or question.get("sku") or "")
            question = {**question, "product_name": _sku_label(catalog, sku)}

            try:
                answer = draft_answer(question)
            except Exception as e:
                errors.append({"question_id": question_id, "stage": "draft",
                               "error": str(e)})
                continue

            answer_text = (answer.text or "").strip()
            if not answer_text:
                errors.append({"question_id": question_id, "stage": "draft",
                               "error": "empty answer text"})
                continue

            try:
                api.question_answer_create(
                    question_id=question_id,
                    answer_text=answer_text,
                    sku=sku,
                )
                answered.append({
                    "question_id": question_id,
                    "chars": len(answer_text),
                    "question": text,
                    "answer": answer_text,
                    "sku": sku,
                    "product_name": question.get("product_name") or "",
                    "created_at": question.get("created_at") or question.get("question_date") or "",
                })
            except Exception as e:
                errors.append({"question_id": question_id, "stage": "post",
                               "error": str(e)})

    summary = {
        "status": "ok" if not errors else "partial",
        "max_answers": max_answers,
        "answered": len(answered),
        "errors": len(errors),
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    if errors:
        print("errors:")
        print(json.dumps(errors, indent=2, ensure_ascii=False))

    _telegram_autoanswer(summary, errors=errors, answered=answered)

    return 0 if not errors else 1


def _tg_send(token: str, chat_id: str, text: str) -> None:
    import httpx
    httpx.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        json={"chat_id": chat_id, "text": text, "parse_mode": "HTML",
              "disable_web_page_preview": True},
        timeout=15,
    )


def _telegram_autoreply(summary: dict, errors: list[dict],
                        replied: list[dict] | None = None) -> None:
    import html

    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        return

    status = summary.get("status")
    flag = "✅" if status == "ok" else "⚠️"
    replied_count = summary.get("replied", 0)
    text = (
        f"<b>{flag} Ozon Auto-Reply</b>\n\n"
        f"Отвечено: <b>{replied_count}</b> из лимита {summary.get('max_replies')}\n"
        f"Без текста (помечены PROCESSED): <b>{summary.get('no_text_marked')}</b>\n"
        f"Ошибок: <b>{summary.get('errors')}</b>"
    )
    if errors:
        preview = errors[:3]
        text += "\n\nПервые ошибки:\n" + "\n".join(
            f"• {e.get('review_id', '?')} [{e.get('stage')}]: {e.get('error', '')[:100]}"
            for e in preview
        )

    try:
        _tg_send(token, chat_id, text)
    except Exception as e:
        print(f"(telegram notify failed: {e})", file=sys.stderr)

    for item in (replied or []):
        stars = "⭐" * int(item.get("rating") or 0) if item.get("rating") else ""
        product_name = item.get("product_name") or str(item.get("sku") or "")
        published = _format_review_date(item.get("published_at") or "")
        header = stars
        if product_name:
            header += f" · {html.escape(product_name)}"
        if published:
            header += f" · {html.escape(published)}"
        question = html.escape((item.get("question") or "").strip())
        answer = html.escape((item.get("answer") or "").strip())
        # Telegram hard limit is 4096; keep a safety margin for the formatting.
        if len(question) > 1500:
            question = question[:1500] + "…"
        if len(answer) > 1500:
            answer = answer[:1500] + "…"
        msg = (
            f"{header}\n\n"
            f"<b>Отзыв:</b>\n<i>{question or '(пустой)'}</i>\n\n"
            f"<b>Ответ Das Experten:</b>\n{answer}"
        )
        try:
            _tg_send(token, chat_id, msg)
        except Exception as e:
            print(f"(telegram Q/A send failed for {item.get('review_id')}: {e})",
                  file=sys.stderr)


def _telegram_autoanswer(summary: dict, errors: list[dict],
                         answered: list[dict] | None = None) -> None:
    import html

    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        return

    answered_count = summary.get("answered", 0)
    error_count = summary.get("errors", 0)
    if answered_count == 0 and error_count == 0:
        return  # silent when nothing happened

    status = summary.get("status")
    flag = "✅" if status == "ok" else "⚠️"
    text = (
        f"<b>{flag} Ozon Auto-Answer (вопросы)</b>\n\n"
        f"Отвечено: <b>{answered_count}</b> из лимита {summary.get('max_answers')}\n"
        f"Ошибок: <b>{error_count}</b>"
    )
    if errors:
        preview = errors[:3]
        text += "\n\nПервые ошибки:\n" + "\n".join(
            f"• {e.get('question_id', '?')} [{e.get('stage')}]: {e.get('error', '')[:300]}"
            for e in preview
        )

    try:
        _tg_send(token, chat_id, text)
    except Exception as e:
        print(f"(telegram notify failed: {e})", file=sys.stderr)

    for item in (answered or []):
        product_name = item.get("product_name") or str(item.get("sku") or "")
        created = _format_review_date(item.get("created_at") or "")
        header = "❓"
        if product_name:
            header += f" {html.escape(product_name)}"
        if created:
            header += f" · {html.escape(created)}"
        question_text = html.escape((item.get("question") or "").strip())
        answer_text = html.escape((item.get("answer") or "").strip())
        if len(question_text) > 1500:
            question_text = question_text[:1500] + "…"
        if len(answer_text) > 1500:
            answer_text = answer_text[:1500] + "…"
        msg = (
            f"{header}\n\n"
            f"<b>Вопрос:</b>\n<i>{question_text or '(пустой)'}</i>\n\n"
            f"<b>Ответ Das Experten:</b>\n{answer_text}"
        )
        try:
            _tg_send(token, chat_id, msg)
        except Exception as e:
            print(f"(telegram Q/A send failed for {item.get('question_id')}: {e})",
                  file=sys.stderr)


def _format_review_date(raw: str) -> str:
    """Convert Ozon's ISO timestamp like '2026-04-20T10:00:00.000Z' to
    '2026-04-20 10:00 МСК' for display. Falls back to the raw string if
    parsing fails."""
    if not raw:
        return ""
    try:
        s = raw.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        msk = dt + timedelta(hours=3)
        return msk.strftime("%Y-%m-%d %H:%M МСК")
    except Exception:
        return raw


def cmd_telegram_ping(args: argparse.Namespace) -> int:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("Telegram not configured (TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID missing)",
              file=sys.stderr)
        return 1
    text = args.text or (
        "<b>✅ Telegram подключён</b>\n\n"
        "Это тестовое сообщение из <code>Ozon Auto-Reply</code>. "
        "Если ты его видишь — бот и chat_id настроены правильно, "
        "дальше сюда будут прилетать сводки и пары «отзыв + ответ»."
    )
    _tg_send(token, chat_id, text)
    print("telegram: sent ok")
    return 0


def cmd_mark_reviews(args: argparse.Namespace) -> int:
    if not args.review_ids:
        print("no review_ids given")
        return 1
    with OzonSellerClient() as c:
        result = SellerAPI(c).change_status(args.review_ids, args.status)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


def cmd_ping_wb(_: argparse.Namespace) -> int:
    with WBSellerClient() as c:
        data = WBFeedbacksAPI(c).count_unanswered()
    print("wb feedbacks auth ok, count-unanswered:")
    print(json.dumps(data, indent=2, ensure_ascii=False))
    return 0


def cmd_debug_wb_feedback(_: argparse.Namespace) -> int:
    """Dump raw JSON of first unanswered feedback — used to inspect WB field names."""
    with WBSellerClient() as c:
        page = WBFeedbacksAPI(c).feedbacks_list(is_answered=False, take=1)
    feedbacks = ((page.get("data") or {}).get("feedbacks")) or []
    if not feedbacks:
        print("No unanswered feedbacks returned")
        return 0
    print(json.dumps(feedbacks[0], indent=2, ensure_ascii=False))
    return 0


def cmd_sync_wb_feedbacks(args: argparse.Namespace) -> int:
    wb_db.init_schema()
    with WBSellerClient() as c:
        result = wb_etl.sync_feedbacks(
            c,
            is_answered=args.answered,
            max_feedbacks=args.max,
        )
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


def cmd_list_wb_recent(args: argparse.Namespace) -> int:
    with WBSellerClient() as c:
        page = WBFeedbacksAPI(c).feedbacks_list(
            is_answered=args.answered, take=args.count, order="dateDesc",
        )
    feedbacks = ((page.get("data") or {}).get("feedbacks")) or []
    if not feedbacks:
        answered_label = "answered" if args.answered else "unanswered"
        print(f"(no feedbacks found, is_answered={answered_label})")
        return 0
    answered_label = "answered" if args.answered else "unanswered"
    print(f"{len(feedbacks)} most recent WB feedbacks ({answered_label}):")
    print("=" * 90)
    for f in feedbacks:
        fid = f.get("id", "")
        rating = f.get("productValuation", "?")
        prod = f.get("productDetails") or {}
        nm_id = prod.get("nmId", "")
        product_name = prod.get("productName") or prod.get("supplierArticle", "")
        author = (f.get("userName") or "").strip() or "(anonymous)"
        created = (f.get("createdDate") or "")[:19]
        text = (f.get("text") or "").strip().replace("\n", " ")
        preview = text[:120] + ("…" if len(text) > 120 else "")
        if not preview:
            preview = "(no text, rating-only)"
        print(f"id: {fid}")
        print(f"    ⭐ {rating}/5   nmId: {nm_id}   «{product_name}»   by: {author}   on: {created}")
        print(f"    “{preview}”")
        print("-" * 90)
    print()
    print("Pick an id and run: python cli.py draft-wb-reply <id>")
    return 0


def _fetch_wb_feedback(api: WBFeedbacksAPI, feedback_id: str) -> dict | None:
    """WB has no /feedback/info endpoint — stream both answered and unanswered
    lists and find the target by id. Slow for big backlogs but fine for a
    one-off draft; the auto-reply loop doesn't use this.
    """
    target = str(feedback_id)
    for is_answered in (False, True):
        for f in api.feedbacks_iter(is_answered=is_answered, page_size=1000):
            if str(f.get("id") or "") == target:
                return f
    return None


def cmd_draft_wb_reply(args: argparse.Namespace) -> int:
    from wb_seller.replier import draft_reply
    with WBSellerClient() as c:
        api = WBFeedbacksAPI(c)
        feedback = _fetch_wb_feedback(api, args.feedback_id)
    if not feedback:
        print(f"error: feedback {args.feedback_id} not found in either answered or unanswered lists",
              file=sys.stderr)
        return 1
    prod = feedback.get("productDetails") or {}
    print("=" * 70)
    print(f"Feedback {args.feedback_id}")
    print("=" * 70)
    print(f"Author: {feedback.get('userName') or '(anonymous)'}")
    print(f"Rating: {feedback.get('productValuation')}/5   nmId: {prod.get('nmId')}")
    print(f"Product: {prod.get('productName') or prod.get('supplierArticle', '')}")
    print(f"Created: {feedback.get('createdDate')}")
    print("-" * 70)
    if feedback.get("pros"):
        print(f"Достоинства: {feedback.get('pros')}")
    if feedback.get("cons"):
        print(f"Недостатки: {feedback.get('cons')}")
    print(feedback.get("text") or "(empty feedback text)")
    print("=" * 70)
    print()
    draft = draft_reply(feedback)
    print("DRAFT REPLY:")
    print("-" * 70)
    print(draft.text)
    print("-" * 70)
    print(
        f"[{draft.model}] in={draft.input_tokens} out={draft.output_tokens} "
        f"cache_read={draft.cache_read_input_tokens} cache_write={draft.cache_creation_input_tokens} "
        f"stop={draft.stop_reason} chars={len(draft.text)}"
    )
    print()
    print("To post this draft as-is, run:")
    print(f"  python cli.py post-wb-reply {args.feedback_id} \"<paste approved text here>\"")
    return 0


def cmd_post_wb_reply(args: argparse.Namespace) -> int:
    if not args.text.strip():
        print("error: empty reply text", file=sys.stderr)
        return 1
    if args.confirm != "YES":
        print("refusing to post without --confirm YES (safety guard)", file=sys.stderr)
        print("the reply is public and cannot be edited after WB moderation window.", file=sys.stderr)
        return 1
    with WBSellerClient() as c:
        result = WBFeedbacksAPI(c).answer_create(
            feedback_id=args.feedback_id,
            text=args.text,
        )
    print("posted:")
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


def cmd_auto_reply_wb(args: argparse.Namespace) -> int:
    """Reply to unanswered WB feedbacks one at a time: stream through
    unanswered, skip rating-only (no text) feedbacks (we can't answer what
    isn't there), and for each feedback with text — draft via Claude and
    POST the answer. Stops after --max-replies successfully posted replies.

    Default: 1 reply per run, so 10-minute cron = up to 6 per hour.
    To clear a backlog, dispatch manually with a larger max-replies.
    """
    from wb_seller.replier import draft_reply

    max_replies = max(1, int(args.max_replies))
    # Hard cap on how many feedbacks we SCAN per run even if most are
    # rating-only / can't-reply — protects us from infinite pagination.
    max_inspect = max_replies * 100
    dry_run = bool(getattr(args, "dry_run", False))
    wb_db.init_schema()

    mode = "DRY-RUN" if dry_run else "LIVE"
    print(f"[auto-reply-wb] start ({mode}) max_replies={max_replies} "
          f"max_inspect={max_inspect}", flush=True)

    replied: list[dict] = []
    rating_only_skipped: list[str] = []
    errors: list[dict] = []
    inspected = 0

    with WBSellerClient() as c:
        api = WBFeedbacksAPI(c)

        # Upfront sanity ping — if token is bad or WB is down, we fail here
        # in seconds rather than hanging later.
        try:
            count = api.count_unanswered() or {}
            data = count.get("data") or {}
            total = data.get("countUnanswered", "?")
            today = data.get("countUnansweredToday", "?")
            print(f"[auto-reply-wb] count-unanswered ok: total={total} today={today}",
                  flush=True)
        except Exception as e:
            print(f"[auto-reply-wb] count-unanswered FAILED: {e}", flush=True)
            raise

        for feedback in api.feedbacks_iter(is_answered=False, page_size=100):
            if len(replied) >= max_replies:
                break
            if inspected >= max_inspect:
                print(f"[auto-reply-wb] reached max_inspect={max_inspect}, stopping",
                      flush=True)
                break
            inspected += 1

            feedback_id = str(feedback.get("id") or "").strip()
            if not feedback_id:
                errors.append({"stage": "validate", "error": "feedback missing id"})
                continue

            text = (feedback.get("text") or "").strip()
            pros = (feedback.get("pros") or "").strip()
            cons = (feedback.get("cons") or "").strip()
            rating = feedback.get("productValuation")
            # WB: treat feedback as "no meaningful text" only if all three
            # free-text fields are empty. Pros/cons often contain the real
            # content even when `text` itself is blank.
            if not text and not pros and not cons:
                rating_only_skipped.append(feedback_id)
                if inspected <= 5 or inspected % 50 == 0:
                    print(f"[{inspected}] skip {feedback_id}: rating-only ({rating}★)",
                          flush=True)
                continue

            prod = feedback.get("productDetails") or {}
            product_name = prod.get("productName") or prod.get("supplierArticle") or "?"
            content_len = len(text) + len(pros) + len(cons)
            print(f"[{inspected}] feedback {feedback_id}: {rating}★ "
                  f"«{product_name[:40]}» content={content_len} chars — drafting via Claude...",
                  flush=True)

            try:
                draft = draft_reply(feedback)
            except Exception as e:
                print(f"  ✗ draft failed: {e}", flush=True)
                errors.append({"feedback_id": feedback_id, "stage": "draft",
                               "error": str(e)})
                continue

            draft_text = (draft.text or "").strip()
            print(f"  draft ready: {len(draft_text)} chars, "
                  f"in={draft.input_tokens} out={draft.output_tokens} "
                  f"cache_read={draft.cache_read_input_tokens}",
                  flush=True)
            if not draft_text:
                errors.append({"feedback_id": feedback_id, "stage": "draft",
                               "error": "empty draft text"})
                continue

            reply_item = {
                "feedback_id": feedback_id,
                "chars": len(draft_text),
                "question": _compose_wb_question(text, pros, cons),
                "answer": draft_text,
                "rating": rating,
                "author": (feedback.get("userName") or "").strip(),
                "nm_id": prod.get("nmId"),
                "product_name": prod.get("productName") or prod.get("supplierArticle") or "",
                "created_at": feedback.get("createdDate") or "",
            }

            if dry_run:
                print(f"  [DRY-RUN] skipping POST — would send "
                      f"{len(draft_text)} chars to feedback {feedback_id}",
                      flush=True)
                reply_item["dry_run"] = True
                replied.append(reply_item)
                _telegram_wb_reply_pair(reply_item)
                continue

            print(f"  POST /feedbacks/answer for {feedback_id}...", flush=True)
            try:
                post_result = api.answer_create(feedback_id=feedback_id, text=draft_text)
                # Log shape of response for the first successful post so we can
                # see exactly what WB returns — helps diagnose endpoint/contract
                # issues. Formatted without curly braces to keep stdout parseable.
                if len(replied) == 0:
                    print("  WB response: " + _fmt_wb_response(post_result),
                          flush=True)
                print(f"  ✓ posted", flush=True)
                replied.append(reply_item)
                # Send to Telegram IMMEDIATELY, not at the end. Invariant: if
                # you see a pair in TG, it's already published on WB. If the
                # workflow gets killed mid-run, we still have a per-reply trail
                # of exactly what went out.
                _telegram_wb_reply_pair(reply_item)
            except Exception as e:
                print(f"  ✗ POST failed: {e}", flush=True)
                errors.append({"feedback_id": feedback_id, "stage": "post",
                               "error": str(e)})

    summary = {
        "status": "ok" if not errors else "partial",
        "max_replies": max_replies,
        "inspected": inspected,
        "replied": len(replied),
        "rating_only_skipped": len(rating_only_skipped),
        "errors": len(errors),
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False), flush=True)
    if errors:
        print("errors:", flush=True)
        print(json.dumps(errors, indent=2, ensure_ascii=False), flush=True)

    _telegram_autoreply_wb(summary, errors=errors)

    return 0 if not errors else 1


def _compose_wb_question(text: str, pros: str, cons: str) -> str:
    """Concatenate WB's three review text fields into one blob for Telegram."""
    parts = []
    if text:
        parts.append(text)
    if pros:
        parts.append(f"[Достоинства] {pros}")
    if cons:
        parts.append(f"[Недостатки] {cons}")
    return "\n".join(parts)


def _fmt_wb_response(resp: dict | None) -> str:
    """Format WB response dict as key=value pairs for a single log line,
    without `{}` so it doesn't collide with the summary-JSON that comes later.
    """
    if not resp:
        return "<empty>"
    pairs = []
    for k, v in resp.items():
        if isinstance(v, str):
            pairs.append(f"{k}=\"{v[:80]}\"")
        elif isinstance(v, (int, float, bool)) or v is None:
            pairs.append(f"{k}={v}")
        else:
            pairs.append(f"{k}=<{type(v).__name__}>")
    return " ".join(pairs)


def _telegram_wb_reply_pair(item: dict) -> None:
    """Send ONE pair (feedback + our reply) to Telegram. Called inline right
    after each successful POST so that if the workflow dies mid-run, we still
    have visibility on whatever was already published to WB."""
    import html

    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        return

    stars = "⭐" * int(item.get("rating") or 0) if item.get("rating") else ""
    product_name = item.get("product_name") or ""
    if not product_name and item.get("nm_id"):
        product_name = f"nmId {item.get('nm_id')}"
    created = _format_review_date(item.get("created_at") or "")
    prefix = "[DRY-RUN] " if item.get("dry_run") else ""
    header = prefix + stars
    if product_name:
        header += f" · {html.escape(product_name)}"
    if created:
        header += f" · {html.escape(created)}"
    question = html.escape((item.get("question") or "").strip())
    answer = html.escape((item.get("answer") or "").strip())
    if len(question) > 1500:
        question = question[:1500] + "…"
    if len(answer) > 1500:
        answer = answer[:1500] + "…"
    msg = (
        f"{header}\n\n"
        f"<b>Отзыв (WB):</b>\n<i>{question or '(пустой)'}</i>\n\n"
        f"<b>Ответ Das Experten:</b>\n{answer}"
    )
    try:
        _tg_send(token, chat_id, msg)
    except Exception as e:
        print(f"(telegram Q/A send failed for {item.get('feedback_id')}: {e})",
              file=sys.stderr)


def _telegram_autoreply_wb(summary: dict, errors: list[dict]) -> None:
    """Send the final summary message. Per-reply pairs are sent inline during
    the loop via _telegram_wb_reply_pair, so this function no longer needs
    the `replied` list."""
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        return

    status = summary.get("status")
    flag = "✅" if status == "ok" else "⚠️"
    replied_count = summary.get("replied", 0)
    text = (
        f"<b>{flag} Wildberries Auto-Reply</b>\n\n"
        f"Отвечено: <b>{replied_count}</b> из лимита {summary.get('max_replies')}\n"
        f"Без текста (пропущены): <b>{summary.get('rating_only_skipped')}</b>\n"
        f"Ошибок: <b>{summary.get('errors')}</b>"
    )
    if errors:
        preview = errors[:3]
        text += "\n\nПервые ошибки:\n" + "\n".join(
            f"• {e.get('feedback_id', '?')} [{e.get('stage')}]: {e.get('error', '')[:100]}"
            for e in preview
        )

    try:
        _tg_send(token, chat_id, text)
    except Exception as e:
        print(f"(telegram notify failed: {e})", file=sys.stderr)


def cmd_ping(_: argparse.Namespace) -> int:
    with OzonPerformanceClient() as c:
        c._auth_header()
        print("auth ok, token cached")
    return 0


def cmd_sync_campaigns(_: argparse.Namespace) -> int:
    db.init_schema()
    with OzonPerformanceClient() as c:
        n = etl.sync_campaigns(c)
    print(f"campaigns upserted: {n}")
    return 0


def cmd_sync_daily(args: argparse.Namespace) -> int:
    db.init_schema()
    if args.days:
        date_to = date.today() - timedelta(days=1)
        date_from = date_to - timedelta(days=args.days - 1)
    else:
        date_from = _parse_date(args.date_from)
        date_to = _parse_date(args.date_to)
    with OzonPerformanceClient() as c:
        n = etl.sync_daily_stats(c, date_from, date_to, args.campaigns or None)
    print(f"daily rows upserted: {n} ({date_from}..{date_to})")
    return 0


def cmd_sync_sku(args: argparse.Namespace) -> int:
    db.init_schema()
    if args.days:
        date_to = date.today() - timedelta(days=1)
        date_from = date_to - timedelta(days=args.days - 1)
    else:
        date_from = _parse_date(args.date_from)
        date_to = _parse_date(args.date_to)
    with OzonPerformanceClient() as c:
        n = etl.sync_sku_stats(c, date_from, date_to, args.campaigns or None,
                               group_by=args.group_by)
    print(f"SKU rows upserted: {n} ({date_from}..{date_to})")
    return 0


def cmd_sync_all(args: argparse.Namespace) -> int:
    db.init_schema()
    with OzonPerformanceClient() as c:
        result = etl.sync_last_n_days(c, days=args.days)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


def cmd_kpi(args: argparse.Namespace) -> int:
    date_to = _parse_date(args.date_to) if args.date_to else date.today() - timedelta(days=1)
    date_from = _parse_date(args.date_from) if args.date_from else date_to - timedelta(days=6)

    totals = analyze.totals(date_from, date_to)
    print(f"--- totals {date_from}..{date_to} ---")
    print(json.dumps(totals, indent=2, ensure_ascii=False, default=str))

    print(f"\n--- top {args.limit} campaigns by spend ---")
    rows = analyze.kpi_by_campaign(date_from, date_to, limit=args.limit)
    _print_table(rows)

    if args.sku:
        print(f"\n--- top {args.limit} SKUs by spend ---")
        rows = analyze.kpi_by_sku(date_from, date_to, limit=args.limit)
        _print_table(rows)
    return 0


def _print_table(rows: list[dict]) -> None:
    if not rows:
        print("(no data)")
        return
    headers = list(rows[0].keys())
    widths = {h: max(len(h), *(len(_fmt(r.get(h))) for r in rows)) for h in headers}
    print(" | ".join(h.ljust(widths[h]) for h in headers))
    print("-+-".join("-" * widths[h] for h in headers))
    for r in rows:
        print(" | ".join(_fmt(r.get(h)).ljust(widths[h]) for h in headers))


def _fmt(v) -> str:
    if v is None:
        return "-"
    if isinstance(v, float):
        return f"{v:.4f}" if abs(v) < 1 else f"{v:.2f}"
    return str(v)


def cmd_debug(args: argparse.Namespace) -> int:
    print("=" * 70)
    print("DEBUG: raw Ozon Performance API responses")
    print("=" * 70)

    with OzonPerformanceClient() as c:
        print("\n--- GET /api/client/campaign ---")
        resp = c.request("GET", "/api/client/campaign")
        print("status:", resp.status_code)
        print("body (first 3000 chars):")
        print(resp.text[:3000])

        date_to = date.today() - timedelta(days=1)
        date_from = date_to - timedelta(days=6)

        from ozon_perf.api import PerformanceAPI
        api = PerformanceAPI(c)
        campaigns = api.list_campaigns()
        print(f"\n--- list_campaigns() parsed {len(campaigns)} items ---")
        if campaigns:
            print("first campaign dict:")
            print(json.dumps(campaigns[0], ensure_ascii=False, indent=2, default=str))

        if campaigns:
            first_ids = [str(c.get("id") or c.get("campaignId")) for c in campaigns[:3]]
            first_ids = [x for x in first_ids if x and x != "None"]
            print(f"\n--- GET /api/client/statistics/daily/json for {first_ids} ---")
            try:
                resp = c.request("GET", "/api/client/statistics/daily/json", params={
                    "campaign_ids": first_ids,
                    "date_from": date_from.isoformat(),
                    "date_to": date_to.isoformat(),
                })
                print("status:", resp.status_code)
                print("body (first 3000 chars):")
                print(resp.text[:3000])
            except Exception as e:
                print("error:", e)
    return 0


def cmd_notify_telegram(args: argparse.Namespace) -> int:
    import html
    import httpx

    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("Telegram not configured (TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID), skipping")
        return 0

    repo = os.environ.get("GITHUB_REPOSITORY", "dasexperten/arams-db")
    branch = os.environ.get("GITHUB_REF_NAME", "claude/explore-capabilities-b9Hjh")
    dashboard_url = f"https://raw.githack.com/{repo}/{branch}/samples/dashboard.html"

    if args.status == "success":
        date_to = date.today() - timedelta(days=1)
        date_from = date_to - timedelta(days=args.days - 1)
        totals = analyze.totals(date_from, date_to)

        def fmt_money(v):
            return f"{v:,.0f}".replace(",", " ") if v else "0"
        def fmt_int(v):
            return f"{v:,}".replace(",", " ") if v else "0"
        def fmt_pct(v):
            return f"{v * 100:.2f}%" if v is not None else "—"
        def fmt_roas(v):
            return f"{v:.2f}x" if v is not None else "—"

        drr = totals.get("drr")
        drr_flag = "\U0001f7e2" if drr is not None and drr < 0.3 else "\U0001f534"

        text = (
            f"<b>✅ Ozon Performance обновлён</b>\n\n"
            f"\U0001f4c5 <b>{date_from}…{date_to}</b> ({args.days} дней)\n"
            f"\U0001f441 Показы: <b>{fmt_int(totals.get('views') or 0)}</b>\n"
            f"\U0001f446 Клики: <b>{fmt_int(totals.get('clicks') or 0)}</b>\n"
            f"\U0001f4e6 Заказы: <b>{fmt_int(totals.get('orders') or 0)}</b>\n"
            f"\U0001f4b0 Выручка: <b>{fmt_money(totals.get('revenue'))} ₽</b>\n"
            f"\U0001f4b8 Расход: <b>{fmt_money(totals.get('spent'))} ₽</b>\n"
            f"{drr_flag} ДРР: <b>{fmt_pct(drr)}</b>\n"
            f"\U0001f3af ROAS: <b>{fmt_roas(totals.get('roas'))}</b>\n\n"
            f"<a href=\"{html.escape(dashboard_url)}\">Открыть дашборд</a>"
        )
    else:
        run_url = args.run_url or ""
        text = (
            f"<b>❌ Ozon Performance: ошибка синхронизации</b>\n\n"
            f"Не удалось обновить данные. "
            + (f"<a href=\"{html.escape(run_url)}\">Посмотреть лог</a>" if run_url else "")
        )

    resp = httpx.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        json={
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "HTML",
            "disable_web_page_preview": True,
        },
        timeout=15,
    )
    print("telegram:", resp.status_code, resp.text[:300])
    resp.raise_for_status()
    return 0


def cmd_dashboard(args: argparse.Namespace) -> int:
    from pathlib import Path
    out = Path(args.out)
    if args.demo:
        dashboard.write_demo(out)
        print(f"demo dashboard written: {out.resolve()}")
        return 0
    db.init_schema()
    if args.days:
        date_to = date.today() - timedelta(days=1)
        date_from = date_to - timedelta(days=args.days - 1)
    else:
        date_to = _parse_date(args.date_to) if args.date_to else date.today() - timedelta(days=1)
        date_from = _parse_date(args.date_from) if args.date_from else date_to - timedelta(days=6)
    dashboard.write(out, date_from, date_to, db_path_label=os.environ.get("OZON_PERF_DB_PATH", ""))
    print(f"dashboard written: {out.resolve()} ({date_from}..{date_to})")
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="ozon-perf", description="Ozon Performance API ETL & analytics")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("init", help="Create SQLite schema (perf + seller)").set_defaults(func=cmd_init)
    sub.add_parser("ping", help="Test Performance credentials / token fetch").set_defaults(func=cmd_ping)
    sub.add_parser("ping-seller", help="Test Seller API credentials (calls /v1/review/count)")\
        .set_defaults(func=cmd_ping_seller)
    sub.add_parser("debug-review", help="Dump raw JSON of first review (inspect field names)")\
        .set_defaults(func=cmd_debug_review)
    sub.add_parser("ping-questions", help="Test questions API — print first question")\
        .set_defaults(func=cmd_ping_questions)
    sub.add_parser(
        "sync-product-skus",
        help="Parse ozon_seller/Ozon Products.xlsx and save ozon_sku → offer_id mapping to DB",
    ).set_defaults(func=cmd_sync_product_skus)
    sub.add_parser("sync-campaigns", help="Fetch campaign list").set_defaults(func=cmd_sync_campaigns)

    sr = sub.add_parser("sync-reviews", help="Fetch reviews from Seller API into SQLite")
    sr.add_argument("--status", choices=["UNPROCESSED", "PROCESSED", "ALL"], default="ALL")
    sr.add_argument("--max", type=int, default=None,
                    help="Stop after N reviews (default: all)")
    sr.add_argument("--with-comments", action="store_true",
                    help="Also pull comments for reviews with comments_amount > 0")
    sr.set_defaults(func=cmd_sync_reviews)

    mr = sub.add_parser("mark-reviews", help="Mark reviews PROCESSED / UNPROCESSED (safe write)")
    mr.add_argument("--status", choices=["PROCESSED", "UNPROCESSED"], required=True)
    mr.add_argument("review_ids", nargs="+")
    mr.set_defaults(func=cmd_mark_reviews)

    lr = sub.add_parser(
        "list-recent",
        help="Print the most recent Ozon reviews with id + preview (no DB write, no POST)",
    )
    lr.add_argument("--status", choices=["UNPROCESSED", "PROCESSED", "ALL"], default="UNPROCESSED")
    lr.add_argument("--count", type=int, default=20, help="How many reviews to show (Ozon requires 20-100)")
    lr.set_defaults(func=cmd_list_recent)

    dr = sub.add_parser(
        "draft-reply",
        help="Fetch a review and generate a draft reply via Claude API (no POST)",
    )
    dr.add_argument("review_id")
    dr.set_defaults(func=cmd_draft_reply)

    pr = sub.add_parser(
        "post-reply",
        help="PUBLICLY post an approved reply to Ozon. Requires --confirm YES.",
    )
    pr.add_argument("review_id")
    pr.add_argument("text", help="The exact approved reply text to publish")
    pr.add_argument("--confirm", default="", help='Pass "YES" to acknowledge the reply is public and irreversible')
    pr.add_argument(
        "--keep-unprocessed",
        action="store_true",
        help="Do NOT mark review as processed after posting (default: mark as processed)",
    )
    pr.set_defaults(func=cmd_post_reply)

    ar = sub.add_parser(
        "auto-reply",
        help="END-TO-END: stream UNPROCESSED reviews one by one, draft via Claude, post reply",
    )
    ar.add_argument(
        "--max-replies", type=int, default=1,
        help="How many replies to post in this run (default: 1). "
             "Reviews with no text are marked PROCESSED and don't count.",
    )
    ar.set_defaults(func=cmd_auto_reply)

    aq = sub.add_parser(
        "auto-answer-questions",
        help="END-TO-END: stream UNANSWERED questions, draft answer via Claude, post to Ozon",
    )
    aq.add_argument(
        "--max-answers", type=int, default=1,
        help="How many answers to post in this run (default: 1). "
             "Increase to clear a backlog.",
    )
    aq.set_defaults(func=cmd_auto_answer_questions)

    tp = sub.add_parser(
        "telegram-ping",
        help="Send a single test message to the Telegram chat (for smoke-testing secrets)",
    )
    tp.add_argument("--text", default=None,
                    help="Custom message text (default: preset 'it works' message)")
    tp.set_defaults(func=cmd_telegram_ping)

    sub.add_parser(
        "ping-wb",
        help="Test WB Feedbacks API credentials (calls /feedbacks/count-unanswered)",
    ).set_defaults(func=cmd_ping_wb)

    sub.add_parser(
        "debug-wb-feedback",
        help="Dump raw JSON of first unanswered WB feedback (inspect field names)",
    ).set_defaults(func=cmd_debug_wb_feedback)

    swb = sub.add_parser("sync-wb-feedbacks", help="Fetch WB feedbacks into SQLite")
    swb.add_argument(
        "--answered", action="store_true",
        help="Fetch already-answered feedbacks (default: only unanswered)",
    )
    swb.add_argument("--max", type=int, default=None,
                     help="Stop after N feedbacks (default: all)")
    swb.set_defaults(func=cmd_sync_wb_feedbacks)

    lrwb = sub.add_parser(
        "list-wb-recent",
        help="Print the most recent WB feedbacks with id + preview (no DB write, no POST)",
    )
    lrwb.add_argument("--answered", action="store_true",
                      help="Show already-answered (default: show unanswered)")
    lrwb.add_argument("--count", type=int, default=20, help="How many feedbacks to show (1..5000)")
    lrwb.set_defaults(func=cmd_list_wb_recent)

    drwb = sub.add_parser(
        "draft-wb-reply",
        help="Fetch a WB feedback and generate a draft reply via Claude API (no POST)",
    )
    drwb.add_argument("feedback_id")
    drwb.set_defaults(func=cmd_draft_wb_reply)

    prwb = sub.add_parser(
        "post-wb-reply",
        help="PUBLICLY post an approved reply to WB. Requires --confirm YES.",
    )
    prwb.add_argument("feedback_id")
    prwb.add_argument("text", help="The exact approved reply text to publish")
    prwb.add_argument("--confirm", default="",
                      help='Pass "YES" to acknowledge the reply is public and irreversible')
    prwb.set_defaults(func=cmd_post_wb_reply)

    arwb = sub.add_parser(
        "auto-reply-wb",
        help="END-TO-END: stream unanswered WB feedbacks one by one, draft via Claude, post reply",
    )
    arwb.add_argument(
        "--max-replies", type=int, default=1,
        help="How many replies to post in this run (default: 1). "
             "Rating-only (no text/pros/cons) feedbacks are skipped and don't count.",
    )
    arwb.add_argument(
        "--dry-run", action="store_true",
        help="Go through feedbacks and draft via Claude, but do NOT POST to WB. "
             "Useful for verifying token + backlog + Claude output without side effects.",
    )
    arwb.set_defaults(func=cmd_auto_reply_wb)

    sd = sub.add_parser("sync-daily", help="Fetch daily campaign stats")
    sd.add_argument("--from", dest="date_from")
    sd.add_argument("--to", dest="date_to")
    sd.add_argument("--days", type=int, help="Last N days (ending yesterday)")
    sd.add_argument("--campaigns", nargs="*", help="Campaign IDs (default: all)")
    sd.set_defaults(func=cmd_sync_daily)

    ss = sub.add_parser("sync-sku", help="Fetch SKU-level stats via async report")
    ss.add_argument("--from", dest="date_from")
    ss.add_argument("--to", dest="date_to")
    ss.add_argument("--days", type=int, help="Last N days (ending yesterday)")
    ss.add_argument("--campaigns", nargs="*", help="Campaign IDs (default: all)")
    ss.add_argument("--group-by", default="DATE", help="DATE | NO_GROUP_BY | PLACEMENT")
    ss.set_defaults(func=cmd_sync_sku)

    sa = sub.add_parser("sync-all", help="Sync campaigns + last N days of stats")
    sa.add_argument("--days", type=int, default=7)
    sa.set_defaults(func=cmd_sync_all)

    kpi = sub.add_parser("kpi", help="Print KPIs by campaign and SKU")
    kpi.add_argument("--from", dest="date_from")
    kpi.add_argument("--to", dest="date_to")
    kpi.add_argument("--limit", type=int, default=20)
    kpi.add_argument("--sku", action="store_true")
    kpi.set_defaults(func=cmd_kpi)

    sub.add_parser("debug", help="Print raw Ozon API responses (no secrets leaked)")\
        .set_defaults(func=cmd_debug)

    tg = sub.add_parser("notify-telegram", help="Send summary/failure to Telegram")
    tg.add_argument("--status", choices=["success", "failure"], default="success")
    tg.add_argument("--days", type=int, default=7)
    tg.add_argument("--run-url", default=None)
    tg.set_defaults(func=cmd_notify_telegram)

    dash = sub.add_parser("dashboard", help="Generate HTML dashboard with charts")
    dash.add_argument("--from", dest="date_from")
    dash.add_argument("--to", dest="date_to")
    dash.add_argument("--days", type=int, help="Last N days (ending yesterday)")
    dash.add_argument("--out", default="dashboard.html")
    dash.add_argument("--demo", action="store_true", help="Use synthetic data (no DB needed)")
    dash.set_defaults(func=cmd_dashboard)

    return p


def main(argv: list[str] | None = None) -> int:
    load_dotenv()
    args = build_parser().parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
